#!/usr/bin/env python3
"""Refresh the public discovery index without republishing article content."""

from __future__ import annotations

import email.utils
import csv
import hashlib
import html
import json
import re
import sys
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
LATEST = ROOT / "data" / "latest.json"
ARCHIVE = ROOT / "data" / "archive.json"
ARCHIVE_CSV = ROOT / "data" / "archive-metadata.csv"
ARCHIVE_XLSX = ROOT / "downloads" / "rasuwa-flood-source-metadata.xlsx"
START = datetime(2026, 8, 26, tzinfo=timezone.utc)
USER_AGENT = "RasuwaFloodSourceIndex/1.0 (+https://rasuwaflood.org/)"
QUERIES = (
    '"Rasuwa flood" OR "Bhote Koshi flood" OR "Bhotekoshi flood" when:30d',
    '"Kyirong Rasuwa flood" OR "Lhende flood" when:30d',
    '(Rasuwa OR Rasuwagadhi) (flood OR avalanche) report assessment satellite when:30d',
)
REQUIRED = re.compile(r"\b(rasuwa|rasuwagadhi|bhote\s*koshi|bhotekoshi|kyirong|lhende)\b", re.I)
HAZARD = re.compile(r"\b(flood|avalanche|landslide|glacier|disaster|mudflow|rockflow)\b", re.I)
TAG = re.compile(r"<[^>]+>")
ARCHIVE_COLUMNS = ("record_id", "date", "first_seen_at", "source", "title", "retained_excerpt", "url", "archive_lookup_url", "content_retention", "rights_note")


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=25) as response:
        return response.read()


def clean(value: str | None, limit: int = 260) -> str:
    text = html.unescape(TAG.sub(" ", value or ""))
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit].rstrip()


def iso_date(value: str | None) -> tuple[str, float]:
    try:
        parsed = email.utils.parsedate_to_datetime(value or "").astimezone(timezone.utc)
    except (TypeError, ValueError, OverflowError):
        parsed = datetime.now(timezone.utc)
    return parsed.date().isoformat(), parsed.timestamp()


def canonical(url: str) -> str:
    parsed = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
    query = [(k, v) for k, v in query if not k.lower().startswith("utm_")]
    return urllib.parse.urlunsplit((parsed.scheme, parsed.netloc.lower(), parsed.path, urllib.parse.urlencode(query), ""))


def google_news() -> list[dict]:
    found: list[dict] = []
    for query in QUERIES:
        url = "https://news.google.com/rss/search?" + urllib.parse.urlencode(
            {"q": query, "hl": "en", "gl": "US", "ceid": "US:en"}
        )
        root = ET.fromstring(fetch(url))
        for node in root.findall("./channel/item"):
            title = clean(node.findtext("title"), 180)
            summary = clean(node.findtext("description"))
            searchable = f"{title} {summary}"
            if not (REQUIRED.search(searchable) and HAZARD.search(searchable)):
                continue
            date, timestamp = iso_date(node.findtext("pubDate"))
            if timestamp < START.timestamp():
                continue
            source_node = node.find("source")
            source = clean(source_node.text if source_node is not None else "News source", 80)
            found.append({
                "source": source,
                "date": date,
                "timestamp": timestamp,
                "title": title,
                "summary": summary or "Relevant reporting discovered by the automated source index.",
                "url": canonical(node.findtext("link") or ""),
                "kind": "News and reporting",
            })
    return found


def previous_items() -> list[dict]:
    items: list[dict] = []
    for path in (ARCHIVE, LATEST):
        try:
            data = json.loads(path.read_text())
            items.extend(item for item in data.get("items", []) if item.get("url"))
        except (FileNotFoundError, json.JSONDecodeError):
            pass
    return items


def write_workbook(items: list[dict], generated_at: str) -> None:
    """Write a human-readable XLSX mirror of the public metadata archive."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read Me"
    archive = workbook.create_sheet("Archive Metadata")
    dark, green, pale, sand = "233B31", "355747", "E9EEE9", "F3EBDD"

    readme.merge_cells("A1:J2")
    readme["A1"] = "Rasuwa Flood — Source Metadata Archive"
    readme["A1"].fill = PatternFill("solid", fgColor=dark)
    readme["A1"].font = Font(bold=True, color="FFFFFF", size=22)
    readme["A1"].alignment = Alignment(vertical="center")
    facts = [
        ("Archive generated", generated_at),
        ("Retained records", len(items)),
        ("Update frequency", "Every three hours"),
        ("Public website", "https://rasuwaflood.org/archive.html"),
        ("Retention policy", "Metadata and source-provided excerpts are retained. Full copyrighted article text is not republished without permission or an open licence."),
    ]
    for row, (label, value) in enumerate(facts, start=4):
        readme.cell(row, 1, label).font = Font(bold=True, color=green)
        readme.cell(row, 1).fill = PatternFill("solid", fgColor=pale)
        readme.cell(row, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 90
    readme.merge_cells("A10:J11")
    readme["A10"] = "Fields include publication date, first-indexed timestamp, publisher, headline, retained excerpt, original URL, archived-copy lookup, retention status and rights note."
    readme["A10"].fill = PatternFill("solid", fgColor=sand)
    readme["A10"].font = Font(italic=True, color="5F4B32")
    readme["A10"].alignment = Alignment(wrap_text=True, vertical="center")
    readme.sheet_view.showGridLines = False

    headers = ("Record ID", "Published date", "First indexed (UTC)", "Publisher", "Headline", "Retained text / excerpt", "Original URL", "Archived-copy lookup", "Content retention", "Rights note")
    archive.merge_cells("A1:J2")
    archive["A1"] = "Rasuwa Flood source metadata"
    archive["A1"].fill = PatternFill("solid", fgColor=dark)
    archive["A1"].font = Font(bold=True, color="FFFFFF", size=20)
    archive.merge_cells("A3:J3")
    archive["A3"] = f"Generated {generated_at} · {len(items)} records · original URLs remain the authoritative source"
    archive["A3"].fill = PatternFill("solid", fgColor=pale)
    archive["A3"].font = Font(italic=True, color=green)
    for column, header in enumerate(headers, start=1):
        cell = archive.cell(5, column, header)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_number, item in enumerate(items, start=6):
        for column, key in enumerate(ARCHIVE_COLUMNS, start=1):
            value = item.get(key, "")
            cell = archive.cell(row_number, column, value)
            cell.alignment = Alignment(wrap_text=column in (4, 5, 6, 9, 10), vertical="top")
        archive.row_dimensions[row_number].height = 48
    widths = (19, 15, 22, 24, 44, 56, 38, 38, 32, 42)
    for column, width in enumerate(widths, start=1):
        archive.column_dimensions[get_column_letter(column)].width = width
    archive.freeze_panes = "C6"
    archive.auto_filter.ref = f"A5:J{len(items) + 5}"
    archive.sheet_view.showGridLines = False
    archive.row_dimensions[5].height = 34
    ARCHIVE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(ARCHIVE_XLSX)


def main() -> int:
    items = previous_items()
    errors: list[str] = []
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    try:
        items.extend(google_news())
    except Exception as exc:  # keep the last good index if a feed is temporarily unavailable
        errors.append(f"Google News RSS: {exc}")

    unique: dict[str, dict] = {}
    for item in items:
        url = canonical(str(item.get("url", "")))
        if not url.startswith(("http://", "https://")):
            continue
        item["url"] = url
        item.setdefault("timestamp", 0)
        item.setdefault("kind", "Source")
        item.setdefault("record_id", hashlib.sha256(url.encode("utf-8")).hexdigest()[:16])
        item.setdefault("first_seen_at", generated_at)
        item["retained_excerpt"] = clean(str(item.get("summary", "")), 500)
        item["archive_lookup_url"] = f"https://web.archive.org/web/*/{url}"
        item["content_retention"] = "Metadata and source-provided excerpt retained"
        item["rights_note"] = "Full text is not republished unless reuse permission or an open licence is documented."
        if url in unique:
            item["first_seen_at"] = unique[url].get("first_seen_at", item["first_seen_at"])
        unique[url] = item

    ordered = sorted(unique.values(), key=lambda item: (item.get("timestamp", 0), item.get("date", "")), reverse=True)
    for item in ordered:
        item.pop("timestamp", None)
    shared = {
        "generated_at": generated_at,
        "collection_method": "Automated metadata and source-provided excerpt discovery; inclusion is not verification. Full copyrighted article text is not republished.",
        "errors": errors,
    }
    latest_payload = {**shared, "count": min(60, len(ordered)), "total_archived": len(ordered), "items": ordered[:60]}
    archive_payload = {**shared, "count": len(ordered), "items": ordered}
    LATEST.parent.mkdir(parents=True, exist_ok=True)
    LATEST.write_text(json.dumps(latest_payload, ensure_ascii=False, indent=2) + "\n")
    ARCHIVE.write_text(json.dumps(archive_payload, ensure_ascii=False, indent=2) + "\n")
    with ARCHIVE_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ARCHIVE_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(ordered)
    try:
        write_workbook(ordered, generated_at)
    except ImportError:
        print("openpyxl unavailable; skipped XLSX refresh", file=sys.stderr)
    print(f"Wrote {len(latest_payload['items'])} latest and {len(ordered)} archived sources")
    if errors:
        print("; ".join(errors), file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
